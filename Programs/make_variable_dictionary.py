"""
make_variable_dictionary.py - Generate an Excel variable dictionary with
summary statistics for ALL features used across the TFT and LightGBM models.

Sources verified against actual code in Programs/02-08, 10, 23.

Output: Data/analysis/survival_descriptives/variable_dictionary.xlsx
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from collections import OrderedDict

sys.path.insert(0, str(Path(__file__).parent))
from importlib.util import spec_from_file_location, module_from_spec
_spec = spec_from_file_location("utilities", Path(__file__).parent / "00_utilities.py")
_mod = module_from_spec(_spec)
_spec.loader.exec_module(_mod)
globals().update({k: v for k, v in vars(_mod).items() if not k.startswith('_')})


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
def _dynamic_variants(col, col_desc):
    src = "Derived in 23_onset_group_benchmark_enhanced.py"
    return [
        (f"{col}_lag3",      f"{col_desc} - 3-month lag",                          src),
        (f"{col}_lag6",      f"{col_desc} - 6-month lag",                          src),
        (f"{col}_delta3",    f"{col_desc} - 3-month change (current minus lag3)",  src),
        (f"{col}_delta6",    f"{col_desc} - 6-month change (current minus lag6)",  src),
        (f"{col}_rollmax6",  f"{col_desc} - rolling 6-month max (lagged 1 month)", src),
        (f"{col}_rollmean6", f"{col_desc} - rolling 6-month mean (lagged 1 month)",src),
    ]


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
VARIABLE_META = OrderedDict()

def _add(name, desc, src):
    VARIABLE_META[name] = (desc, src)

def _sep(label):
    VARIABLE_META[f"__{label}__"] = (label, "")


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("PANEL SKELETON - from 01_build_drug_universe.py")
_add("ndc_11",                    "11-digit National Drug Code identifier", "FDA NDC Directory (product.txt + package.txt)")
_add("year_month",                "Calendar year-month (YYYY-MM)", "Constructed from study period 2020-01 to 2025-09")
_add("product_ndc",               "9-digit product NDC (5-4 format, without package code)", "FDA NDC Directory (package.txt)")
_add("labeler_code",              "FDA labeler code (first 5 digits of NDC)", "Derived from ndc_11")
_add("LABELERNAME",               "Manufacturer / labeler company name (raw)", "FDA NDC Directory (product.txt)")
_add("NONPROPRIETARYNAME",        "Generic (nonproprietary) drug name (raw)", "FDA NDC Directory (product.txt)")
_add("DOSAGEFORMNAME",            "Dosage form name (raw, e.g. TABLET, INJECTION)", "FDA NDC Directory (product.txt)")
_add("ROUTENAME",                 "Route of administration (raw, e.g. ORAL)", "FDA NDC Directory (product.txt)")
_add("MARKETINGCATEGORYNAME",     "Marketing category (raw: ANDA, NDA, BLA)", "FDA NDC Directory (product.txt)")
_add("APPLICATIONNUMBER",         "FDA application number (raw, e.g. ANDA123456)", "FDA NDC Directory (product.txt)")
_add("SUBSTANCENAME",             "Active ingredient substance name(s), semicolon-delimited (raw)", "FDA NDC Directory (product.txt)")
_add("PHARM_CLASSES",             "Pharmacological classes, comma-delimited with [EPC]/[MoA] tags (raw)", "FDA NDC Directory (product.txt)")
_add("DEASCHEDULE",               "DEA controlled substance schedule (raw)", "FDA NDC Directory (product.txt)")
_add("PRODUCTTYPENAME",           "Product type (raw, e.g. HUMAN PRESCRIPTION DRUG)", "FDA NDC Directory (product.txt)")
_add("ACTIVE_NUMERATOR_STRENGTH", "Active ingredient strength numerator (raw)", "FDA NDC Directory (product.txt)")
_add("ACTIVE_INGRED_UNIT",        "Active ingredient unit of measure (raw)", "FDA NDC Directory (product.txt)")
_add("PROPRIETARYNAME",           "Brand / proprietary drug name (raw)", "FDA NDC Directory (product.txt)")
_add("STARTMARKETINGDATE",        "Start marketing date string (raw, YYYYMMDD)", "FDA NDC Directory (product.txt)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("SHORTAGE OUTCOME - from 02_build_shortage_outcome.py")
_add("shortage",              "Binary: drug is in active shortage (1=yes)", "FDA Drug Shortages database (Drugshortages.csv)")
_add("shortage_start",        "Binary: first month of a new shortage episode (1=onset)", "Derived from Initial Posting Date in Drugshortages.csv")
_add("shortage_end",          "Binary: last month of a shortage episode (1=resolution)", "Derived from Change Date / Date Discontinued in Drugshortages.csv")
_add("months_remaining",      "Months remaining in current shortage episode (continuous)", "Derived from shortage start/end dates in Drugshortages.csv")
_add("episode_duration",      "Total duration of shortage episode in months", "Derived from shortage start/end dates in Drugshortages.csv")
_add("episode_censored",      "Binary: shortage episode still active at data end (1=censored)", "Derived from Status field in Drugshortages.csv")
_add("reason_for_shortage",   "Reported reason for shortage (text)", "Reason for Shortage field in Drugshortages.csv")
_add("therapeutic_category",  "Therapeutic category from shortage listing", "Therapeutic Category field in Drugshortages.csv")
_add("shortage_generic_name", "Generic drug name from shortage listing", "Generic Name field in Drugshortages.csv")
_add("shortage_company",      "Company name from shortage listing", "Company Name field in Drugshortages.csv")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("DRUG CHARACTERISTICS - from 03_drug_characteristics.py")
_add("dosage_form",            "Pharmaceutical dosage form, cleaned (e.g., TABLET, INJECTABLE)", "FDA NDC Directory (product.txt DOSAGEFORMNAME)")
_add("is_injectable",          "Binary: 1 if injectable dosage form", "Derived from DOSAGEFORMNAME in NDC Directory")
_add("route",                  "Route of administration, cleaned (e.g., ORAL, INTRAVENOUS)", "FDA NDC Directory (product.txt ROUTENAME)")
_add("is_intravenous",         "Binary: 1 if intravenous route of administration", "Derived from ROUTENAME in NDC Directory")
_add("marketing_category",     "Marketing category, cleaned (ANDA, NDA, BLA, etc.)", "FDA NDC Directory (product.txt MARKETINGCATEGORYNAME)")
_add("is_generic",             "Binary: 1 if ANDA or NDA AUTHORIZED GENERIC", "Derived from MARKETINGCATEGORYNAME in NDC Directory")
_add("substance_name",         "Active ingredient substance name(s), cleaned", "FDA NDC Directory (product.txt SUBSTANCENAME)")
_add("active_ingredient_count","Number of active pharmaceutical ingredients", "Derived from semicolon count in SUBSTANCENAME")
_add("nonproprietary_name",    "Generic (nonproprietary) drug name, cleaned", "FDA NDC Directory (product.txt NONPROPRIETARYNAME)")
_add("proprietary_name",       "Brand / proprietary drug name, cleaned", "FDA NDC Directory (product.txt PROPRIETARYNAME)")
_add("labeler_name",           "Manufacturer / labeler company name, cleaned", "FDA NDC Directory (product.txt LABELERNAME)")
_add("pharm_class_raw",        "Raw pharmacological class string", "FDA NDC Directory (product.txt PHARM_CLASSES)")
_add("therapeutic_class",      "Primary therapeutic classification (EPC or MoA)", "Derived from PHARM_CLASSES in NDC Directory")
_add("dea_schedule",           "DEA controlled substance schedule (cleaned)", "FDA NDC Directory (product.txt DEASCHEDULE)")
_add("is_controlled",          "Binary: 1 if DEA-scheduled controlled substance", "Derived from DEASCHEDULE in NDC Directory")
_add("start_marketing_date",   "Start marketing date (parsed datetime)", "FDA NDC Directory (product.txt STARTMARKETINGDATE)")
_add("application_number",     "FDA application number string (raw)", "FDA NDC Directory (product.txt APPLICATIONNUMBER)")
_add("appl_no",                "Standardized FDA application number (numeric only)", "Derived from APPLICATIONNUMBER in NDC Directory")
_add("appl_type",              "Application type (ANDA, NDA, BLA)", "Derived from APPLICATIONNUMBER prefix in NDC Directory")
_add("product_type",           "Product type (e.g. HUMAN PRESCRIPTION DRUG)", "FDA NDC Directory (product.txt PRODUCTTYPENAME)")
_add("strength",               "Active ingredient strength", "FDA NDC Directory (product.txt ACTIVE_NUMERATOR_STRENGTH)")
_add("ob_approval_date",       "Orange Book approval date (datetime)", "FDA Orange Book (products.txt Approval_Date)")
_add("te_code",                "Therapeutic Equivalence code", "FDA Orange Book (products.txt TE_Code)")
_add("ob_applicant",           "Orange Book applicant / holder name", "FDA Orange Book (products.txt Applicant_Full_Name)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("MARKET STRUCTURE - from 04_market_structure.py")
_add("n_manufacturers",          "Number of active manufacturers for same product", "FDA NDC Directory + historical OB snapshots (2019-2025)")
_add("n_applications",           "Number of ANDA/NDA applications for same product", "FDA Orange Book (products.txt, historical snapshots 2019-2025)")
_add("sole_source",              "Binary: 1 if only one manufacturer", "Derived from n_manufacturers")
_add("recent_generic_entry",     "Binary: new generic competitor entered in recent months", "Derived from NDC Directory + OB snapshot comparison")
_add("recent_manufacturer_exit", "Binary: a manufacturer exited this product recently", "Derived from NDC Directory + OB snapshot comparison")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("MERGERS - from 04b_mergers.py")
_add("ownership_change_12m",   "Binary: ownership change for labeler in past 12 months", "Generic mergers project (ownership_changes_all.csv)")
_add("ownership_change_24m",   "Binary: ownership change for labeler in past 24 months", "Generic mergers project (ownership_changes_all.csv)")
_add("external_merger_12m",    "Binary: external M&A event affecting labeler in past 12 months", "Generic mergers project (ownership_changes_all.csv, is_external flag)")
_add("external_merger_24m",    "Binary: external M&A event affecting labeler in past 24 months", "Generic mergers project (ownership_changes_all.csv, is_external flag)")
_add("n_ownership_changes_24m","Count of ownership changes in past 24 months", "Generic mergers project (ownership_changes_all.csv)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("PRICING - from 04c_pricing.py")
_add("nadac_per_unit",         "National Average Drug Acquisition Cost per unit ($)", "CMS NADAC (Raw Data/NADAC/nadac_*.csv)")
_add("nadac_pct_change_3m",    "NADAC percent change over 3 months", "Derived from CMS NADAC time series")
_add("nadac_pct_change_12m",   "NADAC percent change over 12 months", "Derived from CMS NADAC time series")
_add("nadac_generic_ratio",    "Ratio of NDC price to generic product median", "Derived from CMS NADAC (Corresponding_Generic_Drug_NADAC_Per_Unit)")
_add("nadac_is_low_price",     "Binary: 1 if NADAC price below 25th percentile", "Derived from CMS NADAC distribution")
_add("nadac_vs_market_median", "NADAC relative to product group median price", "Derived from CMS NADAC grouped by product")
_add("nadac_is_observed",      "Binary: 1 if NADAC price directly observed (not imputed/ffill)", "Derived from CMS NADAC match status")
_add("nadac_is_ffill",         "Binary: 1 if NADAC price forward-filled from last observation (6-month limit)", "Derived in 04c_pricing.py")
_add("nadac_is_imputed",       "Binary: 1 if NADAC price imputed at product-level median", "Derived in 04c_pricing.py")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("UTILIZATION - from 04d_utilization.py")
_add("medicaid_rx_count",          "Medicaid prescription count for this NDC", "CMS SDUD (Raw Data/SDUD/sdud_*.csv, Number of Prescriptions)")
_add("medicaid_units",             "Medicaid units reimbursed for this NDC", "CMS SDUD (Raw Data/SDUD/sdud_*.csv, Units Reimbursed)")
_add("medicaid_spending",          "Medicaid total spending for this NDC ($)", "CMS SDUD (Raw Data/SDUD/sdud_*.csv, Total Amount Reimbursed)")
_add("medicaid_rx_trend_4q",       "Medicaid Rx count trend over 4 quarters (% change)", "Derived from CMS SDUD time series")
_add("medicaid_units_trend_4q",    "Medicaid units trend over 4 quarters (% change)", "Derived from CMS SDUD time series")
_add("medicaid_rx_cv_4q",          "Coefficient of variation of Medicaid Rx count over 4 quarters", "Derived from CMS SDUD time series")
_add("partd_total_claims",         "Medicare Part D total claim count", "CMS Part D Drug Spending (medicare_partd_spending_*.csv, Tot_Clms)")
_add("partd_avg_cost_per_claim",   "Medicare Part D average cost per claim ($)", "CMS Part D Drug Spending (Avg_Spnd_Per_Clm)")
_add("partd_total_beneficiaries",  "Medicare Part D total beneficiary count", "CMS Part D Drug Spending (Tot_Benes)")
_add("partd_total_spending",       "Medicare Part D total spending ($)", "CMS Part D Drug Spending (Tot_Spndng)")
_add("has_medicaid_data",          "Missingness indicator: 1 if Medicaid data available", "Derived from medicaid_rx_count presence")
_add("has_partd_data",             "Missingness indicator: 1 if Part D data available", "Derived from partd_total_claims presence")
_add("utilization_product_imputed","Binary: 1 if utilization imputed at product level", "Derived in 04d_utilization.py")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("RECALLS - from 04e_recalls.py")
_add("recall_count_12m",  "Count of FDA recalls for labeler in past 12 months", "openFDA (Raw Data/FDA Recalls/fda_drug_recalls_2018_2025.csv)")
_add("class1_recall_12m", "Count of Class I (most serious) recalls in past 12 months", "openFDA (fda_drug_recalls_2018_2025.csv, classification field)")
_add("recall_count_24m",  "Count of FDA recalls for labeler in past 24 months", "openFDA (fda_drug_recalls_2018_2025.csv)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("INSPECTIONS - from 04f_inspections.py")
_add("oai_inspection_12m",  "Binary: Official Action Indicated inspection in past 12 months", "FDA Inspections (Raw Data/FDA Inspections/, classification field)")
_add("vai_inspection_12m",  "Binary: Voluntary Action Indicated inspection in past 12 months", "FDA Inspections (Raw Data/FDA Inspections/, classification field)")
_add("inspection_count_24m","Total FDA inspections of labeler in past 24 months", "FDA Inspections (Raw Data/FDA Inspections/)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("ADVERSE EVENTS - from 04g_adverse_events.py")
_add("ae_reports_3m",  "Count of adverse event reports in past 3 months", "openFDA FAERS (Raw Data/FDA Adverse Events/ae_counts_by_manufacturer.csv)")
_add("ae_reports_12m", "Count of adverse event reports in past 12 months", "openFDA FAERS (Raw Data/FDA Adverse Events/ae_counts_by_manufacturer.csv)")
_add("ae_trend_12m",   "Year-over-year percent change in adverse event reports", "Derived from openFDA FAERS AE counts")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("API SOURCING - from 04h_api_sourcing.py")
_add("n_api_suppliers",      "Number of registered API suppliers (active Type II DMF holders)", "FDA DMF (Raw Data/FDA DMF/dmf_list_4q2025.xls)")
_add("n_api_countries",      "Number of countries supplying API for this drug", "Derived from DMF holder country inference")
_add("api_india_share",      "Share of API suppliers located in India (0-1)", "Derived from DMF holder country inference")
_add("api_china_share",      "Share of API suppliers located in China (0-1)", "Derived from DMF holder country inference")
_add("api_us_share",         "Share of API suppliers located in US (0-1)", "Derived from DMF holder country inference")
_add("api_india_china_share","Combined share of API suppliers in India + China (0-1)", "Derived from DMF holder country inference")
_add("api_country_hhi",      "Herfindahl-Hirschman Index of API supplier country concentration", "Derived from DMF holder country shares")
_add("has_api_data",          "Missingness indicator: 1 if DMF data matched for this NDC", "Derived from DMF-to-NDC match")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("API DISASTER EXPOSURE - from 04h_api_sourcing.py")
_add("api_disaster_exposure_3m",  "Weighted API supply chain disaster exposure, 3-month window", "EM-DAT (Naturaldisasterdata.xlsx) x FDA DMF country shares")
_add("api_disaster_exposure_12m", "Weighted API supply chain disaster exposure, 12-month window", "EM-DAT (Naturaldisasterdata.xlsx) x FDA DMF country shares")
_add("api_major_disaster_12m",    "Binary: major disaster (>$1B damage) in API source country in past 12 months", "EM-DAT (Naturaldisasterdata.xlsx) x FDA DMF country shares")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("WARNING LETTERS - from 05_warning_letters.py")
_add("warning_letter_6m",  "Count of FDA warning letters to labeler in prior 6 months", "FDA CDER (Raw Data/warning-letters.xlsx)")
_add("warning_letter_12m", "Count of FDA warning letters to labeler in prior 12 months", "FDA CDER (Raw Data/warning-letters.xlsx)")
_add("warning_letter_24m", "Count of FDA warning letters to labeler in prior 24 months", "FDA CDER (Raw Data/warning-letters.xlsx)")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("GEOGRAPHIC & DISASTER - from 06_geographic_disasters.py")
_add("primary_country",   "Primary country of manufacturer (ISO code from facility address)", "FDA facility registrations (Drug Manufacturer locations.zip)")
_add("is_domestic",        "Binary: 1 if manufacturer is US-based", "Derived from facility address country code")
_add("n_facilities",       "Number of registered manufacturing facilities for labeler", "FDA facility registrations (Drug Manufacturer locations.zip)")
_add("n_countries",        "Number of countries with manufacturing facilities", "FDA facility registrations (Drug Manufacturer locations.zip)")
_add("disaster_count_3m",  "Count of natural disasters near facilities in past 3 months", "EM-DAT (Naturaldisasterdata.xlsx) x facility country")
_add("disaster_count_12m", "Count of natural disasters near facilities in past 12 months", "EM-DAT (Naturaldisasterdata.xlsx) x facility country")
_add("major_disaster_12m", "Binary: major disaster (>$1B damage) near facilities in past 12 months", "EM-DAT (Naturaldisasterdata.xlsx) x facility country")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("PATENTS & EXCLUSIVITY - from 07_patents_exclusivity.py")
_add("patent_count",             "Number of active patents listed in Orange Book", "FDA Orange Book (Raw Data/Orange Book Data/patent.txt)")
_add("total_patents_ever",       "Total patents ever listed for this product", "FDA Orange Book (patent.txt)")
_add("months_to_nearest_expiry", "Months until nearest patent expiration (-1 if none)", "Derived from Patent_Expire_Date_Text in patent.txt")
_add("recent_patent_expiry",     "Binary: 1 if a patent expired in last 12 months", "Derived from Patent_Expire_Date_Text in patent.txt")
_add("has_substance_patent",     "Binary: 1 if substance (molecule) patent active", "FDA Orange Book (patent.txt Drug_Substance_Flag)")
_add("has_product_patent",       "Binary: 1 if product (formulation) patent active", "FDA Orange Book (patent.txt Drug_Product_Flag)")
_add("has_active_exclusivity",   "Binary: 1 if any active market exclusivity", "FDA Orange Book (Raw Data/Orange Book Data/exclusivity.txt)")
_add("months_to_exclusivity_end","Months until exclusivity expiration (-1 if none)", "Derived from Exclusivity_Date in exclusivity.txt")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("ASSEMBLY-TIME FEATURES - from 08_assemble_panel.py")
_add("n_repackagers",             "Count of repackager/relabeler NDCs for same drug product", "Derived from LABELERNAME keyword matching in 08_assemble_panel.py")
_add("ob_match_flag",             "Binary: 1 if NDC matched to Orange Book application number", "Derived from appl_no presence in 08_assemble_panel.py")
_add("has_market_structure_data",  "Missingness indicator: 1 if market structure data available", "Derived from n_manufacturers presence")
_add("has_geo_data",              "Missingness indicator: 1 if geographic/facility data available", "Derived from is_domestic presence")
_add("has_patent_data",           "Missingness indicator: 1 if patent data available", "Derived from patent_count/months_to_nearest_expiry presence")
_add("has_merger_data",           "Missingness indicator: 1 if merger data matched", "Derived from ownership_change_12m presence")
_add("has_nadac_data",            "Missingness indicator: 1 if NADAC pricing available", "Derived from nadac_per_unit presence")
_add("has_nadac_trend_3m",        "Missingness indicator: 1 if 3-month price trend calculable", "Derived from nadac_pct_change_3m presence")
_add("has_nadac_trend_12m",       "Missingness indicator: 1 if 12-month price trend calculable", "Derived from nadac_pct_change_12m presence")
_add("has_nadac_market_median",   "Missingness indicator: 1 if market median comparison available", "Derived from nadac_vs_market_median presence")
_add("has_utilization_data",      "Missingness indicator: 1 if any utilization data available", "Derived from has_medicaid_data | has_partd_data")
_add("has_medicaid_trend_data",   "Missingness indicator: 1 if Medicaid trend calculable", "Derived from medicaid_rx_trend_4q presence")
_add("has_warning_letter_data",   "Missingness indicator: 1 if warning letter data available", "Derived from warning_letter_* presence")
_add("has_inspection_data",       "Missingness indicator: 1 if inspection data available", "Derived from oai_inspection_12m/inspection_count_24m presence")
_add("has_recall_data",           "Missingness indicator: 1 if recall data available", "Derived from recall_count_12m/class1_recall_12m presence")
_add("has_adverse_event_data",    "Missingness indicator: 1 if adverse event data available", "Derived from ae_reports_3m/ae_reports_12m/ae_trend_12m presence")
_add("year",                      "Calendar year (integer)", "Derived from year_month[:4]")
_add("month",                     "Calendar month (1-12)", "Derived from year_month[5:7]")
_add("quarter",                   "Calendar quarter (1-4)", "Derived from month")
_add("covid_period",              "Binary: 1 if March 2020 - June 2021", "Constructed from year_month range")
_add("hurricane_season",          "Binary: 1 if June-November", "Constructed from month")
_add("years_on_market",           "Years since first marketing date (continuous)", "Derived from STARTMARKETINGDATE in NDC Directory")
_add("shortage_lag1m",            "Binary: shortage status 1 month prior", "Derived from shortage column, shift(1)")
_add("shortage_lag3m",            "Binary: any shortage in past 3 months", "Derived from shortage column, rolling(3).max()")
_add("shortage_lag12m",           "Binary: any shortage in past 12 months", "Derived from shortage column, rolling(12).max()")
_add("same_ingredient_in_shortage","Binary: another NDC with same active ingredient is in shortage", "Derived from shortage x SUBSTANCENAME groupby in 08_assemble_panel.py")
_add("labeler_shortage_burden",   "Count of other products from same labeler currently in shortage", "Derived from shortage x labeler_code groupby in 08_assemble_panel.py")
_add("time_since_last_shortage",  "Months since last shortage resolution for this NDC (-1 if never)", "Derived from shortage transitions in 08_assemble_panel.py")
_add("labeler_product_count",     "Number of distinct NDCs marketed by same labeler this month", "Derived from labeler_code x year_month groupby in 08_assemble_panel.py")
_add("net_manufacturer_change_12m","Net change in manufacturer count over 12 months", "Derived from n_manufacturers shift(12) in 08_assemble_panel.py")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM ENGINEERED FEATURES - from 23_onset_group_benchmark_enhanced.py")
_add("drug_group_key",       "Drug group identifier: nonproprietary_name | dosage_form", "Derived from nonproprietary_name + dosage_form")
_add("n_ndcs",               "Number of distinct NDCs in this drug group x month", "Derived from ndc_11 nunique per group")
_add("few_manufacturers",    "Binary: 1 if n_manufacturers <= 2", "Derived from n_manufacturers")
_add("few_api_suppliers",    "Binary: 1 if n_api_suppliers <= 2", "Derived from n_api_suppliers")
_add("market_vulnerability", "Composite: sole_source + few_manufacturers + is_injectable", "Derived from panel columns")
_add("quality_signal_any",   "Binary: 1 if any quality signal (warning letter, OAI, VAI, recall, Class I)", "Derived from panel columns")
_add("quality_signal_count", "Count of quality signals (warning letter + OAI + VAI + recall + Class I)", "Derived from panel columns")
_add("disruption_signal",    "Composite: log(disaster_count_12m) + log(api_disaster_12m) + api_major_disaster", "Derived from panel columns")
_add("price_shock",          "Composite: |nadac_pct_change_3m| + 0.5*|nadac_pct_change_12m|", "Derived from panel columns")
_add("supply_chain_risk",    "Composite: api_country_hhi + api_india_china_share + api_major_disaster_12m", "Derived from panel columns")
_add("ingredient_pressure",  "same_ingredient_in_shortage * (1 + few_manufacturers + few_api_suppliers)", "Derived from panel columns")
_add("manufacturer_exit_pressure", "recent_manufacturer_exit + max(0, -net_manufacturer_change_12m)", "Derived from panel columns")
_add("patent_cliff_12m",     "Binary: 1 if patent expires within 12 months", "Derived from months_to_nearest_expiry")
_add("exclusivity_loss_12m", "Binary: 1 if exclusivity expires within 12 months", "Derived from months_to_exclusivity_end")
_add("injectable_sole_source","Binary: is_injectable * sole_source", "Derived from panel columns")
_add("quality_stack",        "Weighted composite: 1.3*quality_signal_count + 0.8*has_inspection + 0.6*ae_trend", "Derived from panel columns")
_add("capacity_stress",      "Composite: (3 - min(n_mfr,3)) + (3 - min(n_api,3)) + log(n_repackagers)", "Derived from panel columns")
_add("contagion_stress",     "Composite: 1.5*same_ingredient_in_shortage + labeler_shortage_burden", "Derived from panel columns")
_add("commercial_risk",      "Composite: |nadac_pct_change_12m| + max(0,-medicaid_rx_trend) + medicaid_rx_cv", "Derived from panel columns")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM GROUP COMPOSITION FEATURES")
_add("n_labelers_group",                    "Number of unique labelers in drug group x month", "Derived from labeler_code nunique per group")
_add("n_application_numbers_group",         "Number of unique application numbers in drug group x month", "Derived from application_number nunique per group")
_add("n_routes_group",                      "Number of unique routes in drug group x month", "Derived from route nunique per group")
_add("n_therapeutic_classes_group",         "Number of unique therapeutic classes in drug group x month", "Derived from therapeutic_class nunique per group")
_add("share_quality_signal_ndcs",           "Share of NDCs in group with any quality signal", "Derived from per-NDC quality flag mean")
_add("share_low_api_ndcs",                  "Share of NDCs in group with n_api_suppliers <= 2", "Derived from per-NDC flag mean")
_add("share_api_concentrated_ndcs",         "Share of NDCs in group with api_country_hhi >= 0.6", "Derived from per-NDC flag mean")
_add("share_india_china_exposed_ndcs",      "Share of NDCs in group with api_india_china_share >= 0.8", "Derived from per-NDC flag mean")
_add("share_low_competition_ndcs",          "Share of NDCs in group with sole_source or n_manufacturers <= 2", "Derived from per-NDC flag mean")
_add("share_price_shock_ndcs",              "Share of NDCs in group with |nadac_pct_change_12m| >= 0.25", "Derived from per-NDC flag mean")
_add("share_ae_spike_ndcs",                 "Share of NDCs in group with ae_trend_12m > 0.25", "Derived from per-NDC flag mean")
_add("share_disaster_risk_ndcs",            "Share of NDCs in group with major_disaster or api_major_disaster", "Derived from per-NDC flag mean")
_add("mean_api_country_hhi_group",          "Mean API country HHI across NDCs in group", "Derived from per-NDC mean")
_add("std_api_country_hhi_group",           "Std dev of API country HHI across NDCs in group", "Derived from per-NDC std")
_add("mean_api_suppliers_group",            "Mean n_api_suppliers across NDCs in group", "Derived from per-NDC mean")
_add("std_api_suppliers_group",             "Std dev of n_api_suppliers across NDCs in group", "Derived from per-NDC std")
_add("mean_manufacturers_group",            "Mean n_manufacturers across NDCs in group", "Derived from per-NDC mean")
_add("std_manufacturers_group",             "Std dev of n_manufacturers across NDCs in group", "Derived from per-NDC std")
_add("mean_abs_nadac_pct_change_12m_group", "Mean |nadac_pct_change_12m| across NDCs in group", "Derived from per-NDC mean")
_add("mean_ae_trend_12m_group",             "Mean ae_trend_12m across NDCs in group", "Derived from per-NDC mean")
_add("dominant_labeler_share_ndcs",         "Share of NDCs held by the dominant labeler in group", "Derived from labeler NDC counts")
_add("dominant_application_share_ndcs",     "Share of NDCs held by the dominant application in group", "Derived from application NDC counts")
_add("ndc_fragmentation_inverse",           "n_ndcs / n_labelers (NDC fragmentation within group)", "Derived from group counts")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM DYNAMIC LAG/DELTA/ROLLING FEATURES (90 total)")

_dynamic_base = [
    ("n_manufacturers",             "Number of active manufacturers"),
    ("n_api_suppliers",             "Number of API suppliers"),
    ("n_facilities",                "Number of manufacturing facilities"),
    ("nadac_per_unit",              "NADAC price per unit"),
    ("nadac_pct_change_3m",         "NADAC 3-month percent change"),
    ("nadac_pct_change_12m",        "NADAC 12-month percent change"),
    ("ae_reports_3m",               "Adverse event reports (3 months)"),
    ("ae_trend_12m",                "Adverse event trend (12 months)"),
    ("labeler_shortage_burden",     "Labeler shortage burden"),
    ("same_ingredient_in_shortage", "Same ingredient in shortage flag"),
    ("disaster_count_3m",           "Disaster count (3 months)"),
    ("api_disaster_exposure_3m",    "API disaster exposure (3 months)"),
    ("quality_signal_count",        "Quality signal count"),
    ("price_shock",                 "Price shock composite"),
    ("commercial_risk",             "Commercial risk composite"),
]
for col, desc in _dynamic_base:
    for name, full_desc, src in _dynamic_variants(col, desc):
        _add(name, full_desc, src)

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM SHORTAGE HISTORY FEATURES")
_add("shortage_months_past_12m",       "Count of shortage months in past 12 months (lagged)", "Derived from shortage column rolling sum")
_add("shortage_months_past_24m",       "Count of shortage months in past 24 months (lagged)", "Derived from shortage column rolling sum")
_add("shortage_starts_past_24m",       "Count of shortage onsets in past 24 months (lagged)", "Derived from shortage_start rolling sum")
_add("shortage_ends_past_24m",         "Count of shortage resolutions in past 24 months (lagged)", "Derived from shortage_end rolling sum")
_add("ever_shortage_before",           "Binary: 1 if any prior shortage onset for this group", "Derived from shortage_start cumsum")
_add("shortage_burden_past_24m",       "Fraction of past 24 months spent in shortage", "Derived: shortage_months_past_24m / 24")
_add("last_group_episode_duration",    "Duration (months) of most recent shortage episode in group", "Derived from shortage state transitions")
_add("mean_group_episode_duration",    "Mean duration (months) of all prior shortage episodes in group", "Derived from shortage state transitions")
_add("max_group_episode_duration",     "Max duration (months) of all prior shortage episodes in group", "Derived from shortage state transitions")
_add("months_since_group_resolution",  "Months since last shortage resolution in group (999 if never)", "Derived from shortage state transitions")
_add("months_since_group_onset",       "Months since last shortage onset in group (999 if never)", "Derived from shortage state transitions")
_add("recent_resolution_rebound",      "Rebound risk: max(0, 6 - months_since_resolution)", "Derived from months_since_group_resolution")
_add("repeat_shortage_flag",           "Binary: 1 if >= 2 shortage onsets in past 24 months", "Derived from shortage_starts_past_24m")
_add("episode_duration_memory",        "Weighted: 0.6*last_episode_duration + 0.4*mean_episode_duration", "Derived from episode duration features")
_add("recurrence_pressure",            "repeat_shortage_flag * (1 + shortage_burden_past_24m)", "Derived from shortage history features")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM RECENCY FEATURES")
_add("months_since_warning",              "Months since last FDA warning letter for group", "Derived from warning_letter_12m > 0 event tracking")
_add("months_since_recall",               "Months since last FDA recall for group", "Derived from recall_count_12m > 0 event tracking")
_add("months_since_oai",                  "Months since last OAI inspection for group", "Derived from oai_inspection_12m > 0 event tracking")
_add("months_since_major_disaster",       "Months since last major disaster affecting group", "Derived from major_disaster_12m > 0 event tracking")
_add("months_since_manufacturer_exit",    "Months since last manufacturer exit from group", "Derived from recent_manufacturer_exit > 0 event tracking")
_add("months_since_ingredient_pressure",  "Months since last same-ingredient shortage pressure", "Derived from same_ingredient_in_shortage > 0 event tracking")
_add("months_since_utilization_observed", "Months since last utilization data observation", "Derived from has_utilization_data > 0 event tracking")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM LAG-SAFE UTILIZATION FEATURES")
_add("medicaid_rx_count_last6",          "Medicaid Rx count, 6-month lag", "Derived from medicaid_rx_count shift(6)")
_add("partd_total_claims_last6",         "Part D total claims, 6-month lag", "Derived from partd_total_claims shift(6)")
_add("medicaid_spending_last6",          "Medicaid spending, 6-month lag", "Derived from medicaid_spending shift(6)")
_add("partd_avg_cost_per_claim_last6",   "Part D avg cost per claim, 6-month lag", "Derived from partd_avg_cost_per_claim shift(6)")
_add("medicaid_trailing_mean_last12",    "Medicaid Rx count trailing 12-month mean (lagged 3 months)", "Derived from medicaid_rx_count shift(3).rolling(12).mean()")
_add("partd_trailing_mean_last12",       "Part D claims trailing 12-month mean (lagged 3 months)", "Derived from partd_total_claims shift(3).rolling(12).mean()")
_add("utilization_claims_ratio_last6",   "Ratio: medicaid_rx_count_last6 / partd_total_claims_last6", "Derived from lagged utilization columns")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM PEER PRESSURE FEATURES")
for cat_col in ["therapeutic_class", "therapeutic_category", "route", "dosage_form"]:
    _add(f"{cat_col}_peer_onset_rate_6m",
         f"Peer onset rate in past 6 months among same {cat_col}", "Derived from shortage_start groupby {0}".format(cat_col))
    _add(f"{cat_col}_peer_shortage_burden_3m",
         f"Peer shortage burden in past 3 months among same {cat_col}", "Derived from shortage groupby {0}".format(cat_col))

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM INTERACTION FEATURES")
_add("peer_vulnerability_interaction",  "therapeutic_class_peer_onset_rate * (1 + market_vulnerability + share_low_competition)", "Derived from peer + engineered features")
_add("quality_peer_interaction",        "route_peer_shortage_burden * (1 + quality_signal_count + share_quality_signal)", "Derived from peer + engineered features")
_add("recurrence_x_peer_pressure",     "recurrence_pressure * (1 + dosage_form_peer_onset_rate)", "Derived from shortage history + peer features")
_add("supplier_concentration_risk",    "dominant_labeler_share + dominant_application_share + share_api_concentrated", "Derived from group composition features")
_add("rebound_risk",                   "recent_resolution_rebound * (1 + episode_duration_memory/12)", "Derived from shortage history features")

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
_sep("LightGBM TARGET")
_add("onset_any6", "Binary: any shortage_start in next 6 months for drug group (target)", "Derived from shortage_start shift(-1 to -6).max()")


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
def _compute_stats(col):
    total = len(col)
    n_miss = int(col.isna().sum())
    row = {"Missingness": f"{n_miss / total:.2%}" if total > 0 else "N/A"}
    if pd.api.types.is_numeric_dtype(col):
        valid = col.dropna()
        if len(valid) > 0:
            row["Min"] = valid.min()
            row["Max"] = valid.max()
            row["Mean"] = valid.mean()
            row["Median"] = valid.median()
            row["SD"] = valid.std()
        else:
            for k in ("Min", "Max", "Mean", "Median", "SD"):
                row[k] = "N/A"
    else:
        row["Min"] = f"{col.nunique()} unique"
        for k in ("Max", "Mean", "Median", "SD"):
            row[k] = ""
    return row


def main():
    print("=" * 70)
    print("Generating variable dictionary with summary statistics")
    print("=" * 70)

    # ---------------------------------------------------------------------------
    panel_path = ANALYSIS / "drug_shortage_panel.parquet"
    panel = None
    try:
        panel = pd.read_parquet(panel_path)
        print(f"  Panel loaded: {panel.shape[0]:,} rows x {panel.shape[1]} cols")
    except OSError as e:
        print(f"  Panel not available: {e}")

    # ---------------------------------------------------------------------------
    rows = []
    panel_cols = set(panel.columns) if panel is not None else set()

    for var_name, (description, source) in VARIABLE_META.items():
        if var_name.startswith("__") and var_name.endswith("__"):
            rows.append({
                "Variable Name": f"-- {description} --",
                "Description": "", "Source": "", "Missingness": "",
                "Min": "", "Max": "", "Mean": "", "Median": "", "SD": "",
            })
            continue

        row = {"Variable Name": var_name, "Description": description, "Source": source}
        if var_name in panel_cols:
            row.update(_compute_stats(panel[var_name]))
        else:
            row["Missingness"] = "Computed at model runtime"
            for k in ("Min", "Max", "Mean", "Median", "SD"):
                row[k] = ""
        rows.append(row)

    # ---------------------------------------------------------------------------
    documented = {k for k in VARIABLE_META if not k.startswith("__")}
    if panel is not None:
        undocumented = sorted(panel_cols - documented)
        if undocumented:
            rows.append({"Variable Name": "-- Additional panel columns (auto-detected) --",
                         "Description": "", "Source": "", "Missingness": "",
                         "Min": "", "Max": "", "Mean": "", "Median": "", "SD": ""})
            for var_name in undocumented:
                row = {"Variable Name": var_name,
                       "Description": "(auto-detected from panel)",
                       "Source": "See source code"}
                row.update(_compute_stats(panel[var_name]))
                rows.append(row)

    df_dict = pd.DataFrame(rows)[["Variable Name", "Description", "Source", "Missingness",
                                   "Min", "Max", "Mean", "Median", "SD"]]

    # Save with the other survival descriptive outputs.
    output_dir = ANALYSIS / "survival_descriptives"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "variable_dictionary.xlsx"
    print(f"\nSaving to {output_path}...")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_dict.to_excel(writer, index=False, sheet_name='Variable Dictionary')
        ws = writer.sheets['Variable Dictionary']

        for ci, cn in enumerate(df_dict.columns, 1):
            ml = max(len(str(cn)), df_dict[cn].astype(str).str.len().max())
            w = {"Description": 80, "Source": 70, "Variable Name": 50}.get(cn, 18)
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = min(ml + 2, w)

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, size=11, color="FFFFFF")
        sep_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))

        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for ri in range(2, ws.max_row + 1):
            is_sep = str(ws.cell(ri, 1).value).startswith("--")
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(ri, ci)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='right' if ci >= 5 else 'left',
                    vertical='top', wrap_text=True)
                if ci >= 5 and isinstance(cell.value, float):
                    if abs(cell.value) >= 1000:
                        cell.number_format = '#,##0.0'
                    elif abs(cell.value) >= 1:
                        cell.number_format = '0.000'
                    else:
                        cell.number_format = '0.0000'
                if is_sep:
                    cell.fill = sep_fill
                    cell.font = Font(bold=True, italic=True)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    n_actual = len([k for k in VARIABLE_META if not k.startswith("__")])
    print(f"\n  Documented variables: {n_actual}")
    print(f"  Total rows in Excel: {len(df_dict)}")
    print(f"  Saved to {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
